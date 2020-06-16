import openpyxl, os

os.chdir('c:\\users\\desktop\\github\\git-repos\\atbs')
workbook = openpyxl.load_workbook('example.xlsx')

'''
type(workbook)
<class 'openpyxl.workbook.workbook.Workbook'>
>>> sheet = workbook.get_sheet_by_name('Sheet1')

Warning (from warnings module):
  File "__main__", line 1
DeprecationWarning: Call to deprecated function get_sheet_by_name (Use wb[sheetname]).
>>> workbook.get_sheet_names()

Warning (from warnings module):
  File "__main__", line 1
DeprecationWarning: Call to deprecated function get_sheet_names (Use wb.sheetnames).
['Sheet1']
>>> wb.sheetnames
Traceback (most recent call last):
  File "<pyshell#12>", line 1, in <module>
    wb.sheetnames
NameError: name 'wb' is not defined
>>> workbook.sheetnames
['Sheet1']
>>> sheet = workbook['Sheet1']
>>> sheet
<Worksheet "Sheet1">
>>> sheet.cell(row=1, column=2)
<Cell 'Sheet1'.B1>
>>> for i in range(1, 8):
	print(i, sheet.cell(row = i, column = 2).value)

	
1 None
2 None
3 None
4 None
5 None
6 None
7 None
'''
